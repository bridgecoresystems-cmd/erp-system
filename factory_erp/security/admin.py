from django.contrib import admin
from django.http import HttpResponse
from django.urls import path
from django.utils.html import format_html
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

from .models import SecurityLog, Shift


@admin.register(SecurityLog)
class SecurityLogAdmin(admin.ModelAdmin):
    list_display = ('employee', 'action', 'timestamp', 'security_guard', 'colored_action')
    list_filter = ('action', 'timestamp', 'security_guard', 'employee__department')
    search_fields = ('employee__name', 'security_guard__username', 'notes')
    readonly_fields = ('timestamp',)
    date_hierarchy = 'timestamp'
    list_per_page = 50
    
    fieldsets = (
        ('Основная информация', {
            'fields': ('employee', 'action', 'timestamp')
        }),
        ('Дополнительно', {
            'fields': ('security_guard', 'notes'),
            'classes': ('collapse',)
        }),
    )
    
    def colored_action(self, obj):
        """Цветное отображение действия"""
        if obj.action == 'in':
            color = '#27ae60'
            icon = '➡️'
        else:
            color = '#e74c3c'
            icon = '⬅️'
        
        return format_html(
            '<span style="color: {}; font-weight: bold;">{} {}</span>',
            color, icon, obj.get_action_display()
        )
    colored_action.short_description = 'Действие'
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('export-excel/', self.export_excel, name='securitylog_export_excel'),
        ]
        return custom_urls + urls
    
    def export_excel(self, request):
        """Экспорт всех логов в Excel"""
        # Получаем все логи
        logs = SecurityLog.objects.select_related('employee', 'security_guard').order_by('-timestamp')
        
        # Создаем Excel файл
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Все логи охраны'
        
        # Стили
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Заголовки
        headers = ['№', 'Дата', 'Время', 'Сотрудник', 'Отдел', 'Действие', 'Охранник', 'Примечания']
        
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Данные
        for row_num, log in enumerate(logs, 2):
            worksheet.cell(row=row_num, column=1, value=row_num - 1)
            worksheet.cell(row=row_num, column=2, value=log.timestamp.strftime('%d.%m.%Y'))
            worksheet.cell(row=row_num, column=3, value=log.timestamp.strftime('%H:%M'))
            worksheet.cell(row=row_num, column=4, value=log.employee.get_full_name())
            worksheet.cell(row=row_num, column=5, value=log.employee.department or '')
            worksheet.cell(row=row_num, column=6, value=log.get_action_display())
            worksheet.cell(row=row_num, column=7, value=log.security_guard.username if log.security_guard else '')
            worksheet.cell(row=row_num, column=8, value=log.notes)
        
        # Автоширина колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Подготовка ответа
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f"all_security_logs_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        workbook.save(response)
        return response
    
    def changelist_view(self, request, extra_context=None):
        """Добавляем кнопку экспорта в админку"""
        extra_context = extra_context or {}
        extra_context['export_url'] = 'export-excel/'
        return super().changelist_view(request, extra_context=extra_context)


@admin.register(Shift)
class ShiftAdmin(admin.ModelAdmin):
    list_display = ('guard', 'start_time', 'end_time', 'duration', 'is_active', 'shift_status')
    list_filter = ('is_active', 'start_time', 'guard')
    search_fields = ('guard__username', 'guard__first_name', 'guard__last_name')
    readonly_fields = ('duration',)
    date_hierarchy = 'start_time'
    
    fieldsets = (
        ('Смена', {
            'fields': ('guard', 'start_time', 'end_time', 'is_active')
        }),
        ('Информация', {
            'fields': ('duration',),
            'classes': ('collapse',)
        }),
    )
    
    def duration(self, obj):
        """Продолжительность смены"""
        if obj.start_time and obj.end_time:
            delta = obj.end_time - obj.start_time
            hours = delta.total_seconds() // 3600
            minutes = (delta.total_seconds() % 3600) // 60
            return f"{int(hours)}ч {int(minutes)}м"
        elif obj.start_time and obj.is_active:
            from django.utils import timezone
            delta = timezone.now() - obj.start_time
            hours = delta.total_seconds() // 3600
            minutes = (delta.total_seconds() % 3600) // 60
            return f"{int(hours)}ч {int(minutes)}м (активна)"
        return "—"
    duration.short_description = 'Продолжительность'
    
    def shift_status(self, obj):
        """Статус смены с цветом"""
        if obj.is_active:
            return format_html(
                '<span style="color: #27ae60; font-weight: bold;">🟢 Активна</span>'
            )
        else:
            return format_html(
                '<span style="color: #e74c3c; font-weight: bold;">🔴 Завершена</span>'
            )
    shift_status.short_description = 'Статус'


# Кастомизация админки
admin.site.site_header = 'Управление безопасностью'
admin.site.site_title = 'Security Admin'
admin.site.index_title = 'Панель управления безопасностью'