from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Q
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from employees.models import Employee
from .models import SecurityLog, Shift

# Добавьте эти функции в security/views.py

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json

@csrf_exempt
def latest_access_api(request):
    """API для получения последнего касания карты"""
    try:
        # Получаем последнюю запись
        latest_log = SecurityLog.objects.select_related('employee').order_by('-timestamp').first()
        
        if latest_log:
            # Проверяем, была ли запись в последние 10 секунд (считаем "новой")
            time_diff = timezone.now() - latest_log.timestamp
            is_new = time_diff.total_seconds() <= 3
            
            employee_data = {
                'full_name': latest_log.employee.get_full_name() if hasattr(latest_log.employee, 'get_full_name') else f"{latest_log.employee.last_name} {latest_log.employee.first_name}",
                'position': latest_log.employee.position or '',
                'department': latest_log.employee.department or '',
                'employee_id': latest_log.employee.employee_id or latest_log.employee.id,
                'rfid_uid': latest_log.employee.rfid_uid or '—',
                'photo_url': latest_log.employee.photo.url if latest_log.employee.photo else '/static/images/no-photo.png'
            }
            
            return JsonResponse({
                'success': True,
                'has_new_access': is_new,
                'access_time': latest_log.timestamp.isoformat(),
                'action': latest_log.action,
                'employee': employee_data
            })
        else:
            return JsonResponse({
                'success': True,
                'has_new_access': False
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@csrf_exempt 
def dashboard_stats_api(request):
    """API для статистики dashboard"""
    try:
        from datetime import date
        
        today = date.today()
        
        # Количество проходов сегодня
        today_accesses = SecurityLog.objects.filter(timestamp__date=today).count()
        
        # Сотрудники в офисе (последнее действие = вход)
        employees_in_office = []
        employees = Employee.objects.filter(is_active=True)
        
        for employee in employees:
            last_log = SecurityLog.objects.filter(employee=employee).order_by('-timestamp').first()
            if last_log and last_log.action == 'in':
                employees_in_office.append(employee)
        
        # Общее количество активных сотрудников
        total_active = Employee.objects.filter(is_active=True).count()
        
        return JsonResponse({
            'success': True,
            'today_accesses': today_accesses,
            'unique_employees_today': len(employees_in_office),
            'total_active_employees': total_active,
            'last_update': timezone.localtime(timezone.now()).strftime('%H:%M:%S')
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@csrf_exempt
def rfid_scan_api(request):
    """API для приема данных от ESP32"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            rfid_uid = data.get('rfid_uid')
            
            if not rfid_uid:
                return JsonResponse({'success': False, 'error': 'RFID UID not provided'})
            
            # Ищем сотрудника по RFID
            try:
                employee = Employee.objects.get(rfid_uid=rfid_uid, is_active=True)
            except Employee.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Employee not found'})
            
            # Определяем действие (вход/выход) по последней записи
            last_log = SecurityLog.objects.filter(employee=employee).order_by('-timestamp').first()
            
            if last_log and last_log.action == 'in':
                action = 'out'  # Если последнее действие было вход, то сейчас выход
            else:
                action = 'in'   # Если последнее действие было выход или записей нет, то вход
            
            # Создаем запись
            SecurityLog.objects.create(
                employee=employee,
                action=action,
                timestamp=timezone.now(),
                notes=f'Автоматическое сканирование RFID: {rfid_uid}'
            )
            
            return JsonResponse({
                'success': True,
                'action': action,
                'employee_name': employee.get_full_name() if hasattr(employee, 'get_full_name') else f"{employee.last_name} {employee.first_name}",
                'message': f'Сотрудник {"вошел" if action == "in" else "вышел"}'
            })
            
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid JSON'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Only POST method allowed'})

def is_security_guard(user):
    """Проверка, является ли пользователь охранником"""
    return (
        user.groups.filter(name='Security').exists() or 
        user.groups.filter(name='Security_Users').exists() or  # Поддержка старой группы
        user.is_superuser
    )


@login_required
@user_passes_test(is_security_guard)
def security_dashboard(request):
    """Панель охраны"""
    employees = Employee.objects.filter(is_active=True).order_by('last_name', 'first_name')
    
    # Последние записи
    recent_logs = SecurityLog.objects.select_related('employee', 'security_guard').order_by('-timestamp')[:10]
    
    # Активная смена
    active_shift = Shift.objects.filter(guard=request.user, is_active=True).first()
    
    # Сотрудники в офисе (последнее действие - вход)
    employees_in_office = []
    for employee in employees:
        last_log = SecurityLog.objects.filter(employee=employee).order_by('-timestamp').first()
        if last_log and last_log.action == 'in':
            employees_in_office.append(employee)
    
    context = {
        'employees': employees,
        'recent_logs': recent_logs,
        'active_shift': active_shift,
        'employees_in_office': employees_in_office,
        'today': timezone.now().date(),
    }
    return render(request, 'security/dashboard.html', context)


@login_required
@user_passes_test(is_security_guard)
def log_action(request, employee_id, action):
    """Логирование входа/выхода сотрудника"""
    employee = get_object_or_404(Employee, id=employee_id)
    
    # Создаем запись
    SecurityLog.objects.create(
        employee=employee,
        action=action,
        security_guard=request.user,
        timestamp=timezone.now()
    )
    
    action_text = "вошел" if action == "in" else "вышел"
    messages.success(request, f"{employee.get_full_name()} {action_text}")
    
    return redirect('security:dashboard')


@login_required
@user_passes_test(is_security_guard)
def start_shift(request):
    """Начало смены"""
    # Завершаем предыдущую смену если есть
    Shift.objects.filter(guard=request.user, is_active=True).update(
        end_time=timezone.now(),
        is_active=False
    )
    
    # Начинаем новую смену
    Shift.objects.create(
        guard=request.user,
        start_time=timezone.now()
    )
    
    messages.success(request, "Смена начата")
    return redirect('security:dashboard')


@login_required
@user_passes_test(is_security_guard)
def end_shift(request):
    """Окончание смены"""
    active_shift = Shift.objects.filter(guard=request.user, is_active=True).first()
    if active_shift:
        active_shift.end_time = timezone.now()
        active_shift.is_active = False
        active_shift.save()
        messages.success(request, "Смена завершена")
    
    return redirect('security:dashboard')


@login_required
@user_passes_test(is_security_guard)
def export_logs_excel(request):
    """Экспорт логов в Excel"""
    # Получаем параметры фильтрации
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    employee_id = request.GET.get('employee')
    
    # Базовый queryset
    logs = SecurityLog.objects.select_related('employee', 'security_guard')
    
    # Применяем фильтры
    if date_from:
        date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
        logs = logs.filter(timestamp__date__gte=date_from)
    else:
        # По умолчанию - за последнюю неделю
        date_from = timezone.now().date() - timedelta(days=7)
        logs = logs.filter(timestamp__date__gte=date_from)
    
    if date_to:
        date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
        logs = logs.filter(timestamp__date__lte=date_to)
    
    if employee_id:
        logs = logs.filter(employee_id=employee_id)
    
    logs = logs.order_by('-timestamp')
    
    # Создаем Excel файл
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Логи охраны'
    
    # Стили
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Заголовки
    headers = ['№', 'Дата', 'Время', 'Сотрудник', 'Действие', 'Охранник', 'Примечания']
    
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
# Данные
    for row_num, log in enumerate(logs, 2):
        worksheet.cell(row=row_num, column=1, value=row_num - 1)
        worksheet.cell(row=row_num, column=2, value=log.timestamp.strftime('%d.%m.%Y'))
        worksheet.cell(row=row_num, column=3, value=log.timestamp.strftime('%H:%M'))
        
        # Безопасное создание полного имени
        full_name = f"{log.employee.last_name or ''} {log.employee.first_name or ''}"
        if log.employee.middle_name:
            full_name += f" {log.employee.middle_name}"
        worksheet.cell(row=row_num, column=4, value=full_name.strip())
        
        worksheet.cell(row=row_num, column=5, value=log.get_action_display())
        worksheet.cell(row=row_num, column=6, value=log.security_guard.username if log.security_guard else '')
        worksheet.cell(row=row_num, column=7, value=log.notes)
    # Автоширина колонок
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Подготовка ответа
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f"security_logs_{date_from.strftime('%Y%m%d')}"
    if date_to:
        filename += f"_to_{date_to.strftime('%Y%m%d')}"
    filename += ".xlsx"
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    workbook.save(response)
    return response


@login_required
@user_passes_test(is_security_guard)
def logs_report(request):
    """Отчет по логам с фильтрацией"""
    # Получаем параметры
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    employee_id = request.GET.get('employee')
    
    # Базовый queryset
    logs = SecurityLog.objects.select_related('employee', 'security_guard')
    
    # Применяем фильтры
    if date_from:
        date_from_parsed = datetime.strptime(date_from, '%Y-%m-%d').date()
        logs = logs.filter(timestamp__date__gte=date_from_parsed)
    
    if date_to:
        date_to_parsed = datetime.strptime(date_to, '%Y-%m-%d').date()
        logs = logs.filter(timestamp__date__lte=date_to_parsed)
    
    if employee_id:
        logs = logs.filter(employee_id=employee_id)
    
    logs = logs.order_by('-timestamp')[:100]  # Лимит для производительности
    
    employees = Employee.objects.filter(is_active=True).order_by('last_name', 'first_name')
    
    context = {
        'logs': logs,
        'employees': employees,
        'date_from': date_from,
        'date_to': date_to,
        'selected_employee': employee_id,
    }
    
    return render(request, 'security/logs_report.html', context)


# ===== AJAX POLLING API ENDPOINTS =====

@login_required
def access_polling_api(request):
    """
    API для AJAX polling - получение последних событий доступа.
    Возвращает последние 50 записей для отображения на мониторе безопасности.
    """
    try:
        # Получаем последние записи
        logs = SecurityLog.objects.select_related('employee').order_by('-timestamp')[:50]
        
        data = []
        for log in logs:
            data.append({
                'id': log.id,
                'employee': {
                    'id': log.employee.id,
                    'full_name': log.employee.get_full_name() if hasattr(log.employee, 'get_full_name') else f"{log.employee.last_name} {log.employee.first_name}",
                    'department': log.employee.department or '',
                    'position': log.employee.position or '',
                    'employee_id': log.employee.employee_id or log.employee.id,
                    'photo_url': log.employee.photo.url if getattr(log.employee, 'photo', None) else '/static/images/no-photo.png'
                },
                'action': log.action,
                'action_display': log.get_action_display(),
                'timestamp': log.timestamp.isoformat(),
                'timestamp_display': log.timestamp.strftime('%H:%M:%S'),
                'notes': log.notes or '',
            })
        
        return JsonResponse({
            'success': True,
            'data': data,
            'count': len(data),
            'timestamp': timezone.now().isoformat()
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)